# -*- coding:utf-8 -*-
__author__ = 'Darphee'




#  Last Modified: Wed Oct 25 15:59:38 2023




import os
import re
import openpyxl
import sys



excel_name= "jwq40260_ipif.xlsx"

def gen_verilog_empty(sheet_number):
    #input_file= os.getcwd()+"\\"+excel_name
    input_file= os.getcwd()+"/"+excel_name

    #1、打开excel的工作表
    wb = openpyxl.load_workbook(input_file)
    #2、获取工作表的名字test,返回的是一个list
    module_name = wb.sheetnames[sheet_number]
    #3、list转为str
    top_name=''.join(module_name)
    #top_name = "dig_top"
    #生成一个test.v，其中写入的是将test.xlsx中的信息转化为verilog的代码框架
    #gen_file= os.getcwd()+"\\"+top_name+'.sv'
    gen_file= os.getcwd()+"/"+top_name+'.sv'
    print(gen_file)
    #每次运行的时候检测test.v是否存在，如果存在则删除
    if os.path.exists(gen_file):
        os.remove(gen_file)
    #print(''.join(module_name))#list ---> str
    #打开excel名字为test的工作表
    table= wb[top_name] #open sheet by name
    #统计表格中的行数和列数
    row_num = table.max_row
    col_num = table.max_column
    ##定义列表变量
    ##端口名字 clk rst_n din dout
    port_name=[]
    ##位宽 1 1 4 3
    connect_name=[]

    port_width=[]##
    ## 端口方向 input output inout
    port_type=[]
    ## 信号复位值
    port_rst =[]
    ## 注释
    port_comment=[]


    for i in range(row_num):
        port_name.append   (table.cell(row=i+1,column=2 ).value)##excel中的第n列
        connect_name.append(table.cell(row=i+1,column=4 ).value)##excel中的第n列
        port_type.append   (table.cell(row=i+1,column=7 ).value)##excel中的第n列
        port_width.append  (table.cell(row=i+1,column=8 ).value)##excel中的第n列
        port_rst.append    (table.cell(row=i+1,column=9 ).value)##excel中的第n列
        port_comment.append(table.cell(row=i+1,column=10).value)##excel中的第n列
        #print(port_name)
        #print(port_type)
        #print(port_width)
        #print(port_rst)

    ##打开test.v并写入从excel信息对应的verilog代码

    with open(gen_file,'a') as file_obj:
        file_obj.write('module '+top_name+' #(\n')
        file_obj.write('    parameter UDLY = 1\n')
        file_obj.write(')(\n')
        for i in range(row_num-1):
            if(port_type[i+1] == "I"):
                port_str="    input "
            elif(port_type[i+1] == "O"):
                port_str='    output'

            if(port_width[i+1]== 1):
                width_str="    "
            else:
                width_str=" ["+str(port_width[i+1]-1)+':0]'

            if(i<row_num-2):
                file_obj.write(port_str+" wire "+width_str +port_name[i+1]+',//' + str(port_comment[i+1])+ ' \n')
            else:
                file_obj.write(port_str+" wire "+ ' [' + str(port_width[i + 1] - 1) + ':0]' + port_name[i + 1] + '//'+ str(port_comment[i+1]) + '\n );\n\n')


        for i in range(row_num-1):
            cell_null = table.cell(row = i+2 , column=4).value 
            if(port_width[i+1]== 1):
                width_str=""
            else:
                width_str=" ["+str(port_width[i+1]-1)+':0]'

            if cell_null is not None:
                if(port_type[i+1] == "I"):
                    file_obj.write("wire   "+width_str +str(connect_name[i+1])+';//'+ str(port_type[i+1]) + '\n')
                    file_obj.write("assign "+ str(connect_name[i+1]) +" = " + str(port_name[i+1]) + ';' + '\n' )
                else:
                    file_obj.write("wire   "+width_str + str(connect_name[i+1])+';\n')
                    file_obj.write("assign "+ str(port_name[i+1]) +" = " + str(connect_name[i+1]) + ';//' + str(port_type[i+1])  + '\n')

            #if(port_type[i+1] == "O"):
            #    file_obj.write("assign "+ port_name[i+1] +" = " + str(port_rst[i+1]) +';\n')
            #else:
            #    file_obj.write('')

        file_obj.write('endmodule')

def main():
    gen_verilog_empty(sheet_number=2)
    gen_verilog_empty(sheet_number=3)
    gen_verilog_empty(sheet_number=4)


if __name__ == "__main__":
    main()

